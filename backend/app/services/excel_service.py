"""Excel export for an approved payroll run.

Layout mirrors excel/payroll_export.xlsx (the bank's bulk-upload template): sheet
"Payroll Export", title block in rows 1-2, headers on row 5, data from row 6, a
bold Total row immediately after the data, then two "Check:" audit rows, and finally
a "Salary Disbursement Summary" table — one row per employee, Phone Number | Amount |
Comment. This table is the bank's actual bulk-upload payload: it is ingested from the
raw cells and NOT recalculated, so Amount is a stored value equal to that employee's
Net Salary above (formulas would read blank), formatted as a plain number with no
"KES" label. Comment is the fixed narration "salary".

Two deliberate differences from that template file:

- Upload-critical numbers — every per-employee amount, and the Total row — are
  written as the run's *stored, computed values*, not as `=C-D-E` / `=SUM(...)`
  formulas. Two reasons: (1) an approved run is immutable financial history, so
  the file must carry exactly what was approved, and (2) openpyxl writes formulas
  with no cached result, so a reader that does not recalculate (a raw bulk-upload
  parser, a headless conversion) would see an empty Net column — the single worst
  failure for a payroll file, since Net is the amount paid. Values are always
  readable regardless of the reader.

- The two "Check:" rows ARE kept as live formulas, exactly as the template. They
  are audit aids meant to be eyeballed in Excel, not upload data, so the no-cache
  caveat is harmless for them — and a formula genuinely re-verifies the file if a
  cell is ever hand-edited, which a hardcoded "OK" could not.

Decimal -> float conversion happens only at the point of writing a cell.
"""
import calendar
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Employee, PayrollLine, PayrollRun
from app.money import to_money

MONEY_FORMAT = '#,##0.00" KES"'
# Bank-upload amounts: plain number, two decimals, no currency label and no thousands
# separator (a comma could trip a raw CSV/cell parser).
PLAIN_AMOUNT_FORMAT = "0.00"

HEADERS = [
    "Employee Name",
    "Phone Number",
    "Gross Salary (KES)",
    "Loan Deduction (KES)",
    "Advance Deduction (KES)",
    "Net Salary (KES)",
]
COLUMN_WIDTHS = {"A": 22, "B": 16, "C": 20, "D": 20, "E": 22, "F": 20}
HEADER_ROW = 5
FIRST_DATA_ROW = 6

_ZERO = Decimal("0.00")


def export_filename(run: PayrollRun) -> str:
    return f"payroll_{run.month:02d}_{run.year}.xlsx"


def build_payroll_export(db: Session, run: PayrollRun) -> BytesIO:
    """Build the workbook in memory and return a rewound byte stream."""
    rows = db.execute(
        select(PayrollLine, Employee.name, Employee.phone)
        .join(Employee, Employee.id == PayrollLine.employee_id)
        .where(PayrollLine.payroll_run_id == run.id)
        .order_by(Employee.name, PayrollLine.id)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Export"

    ws["A1"] = "Monthly Payroll Export"
    ws["A1"].font = Font(bold=True)
    period = f"{calendar.month_name[run.month]} {run.year}"
    approved = run.approved_at.strftime("%Y-%m-%d") if run.approved_at else "-"
    ws["A2"] = f"Payroll for {period} — approved {approved}"

    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=header)
        cell.font = Font(bold=True)

    # Running totals accumulated as Decimals; the Total row is written as values,
    # so there is no SUM-over-an-empty-range edge case when a run has no lines.
    totals = {3: _ZERO, 4: _ZERO, 5: _ZERO, 6: _ZERO}

    # (phone, net) per employee, in table order — feeds the summary below.
    summary_entries: list[tuple[str, Decimal]] = []

    row_num = FIRST_DATA_ROW
    for line, name, phone in rows:
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=phone)
        money = (
            (3, line.gross_salary),
            (4, line.loan_deduction),
            (5, line.advance_deduction),
            (6, line.net_salary),
        )
        for col, value in money:
            cell = ws.cell(row=row_num, column=col, value=float(value))
            cell.number_format = MONEY_FORMAT
            totals[col] += to_money(value)
        summary_entries.append((phone, line.net_salary))
        row_num += 1

    # Total row, immediately after the data (matches the template's layout).
    total_row = row_num
    label = ws.cell(row=total_row, column=1, value="Total")
    label.font = Font(bold=True)
    for col in (3, 4, 5, 6):
        cell = ws.cell(row=total_row, column=col, value=float(totals[col]))
        cell.number_format = MONEY_FORMAT
        cell.font = Font(bold=True)

    # Blank spacer row, then the two audit checks (live formulas, per the module
    # docstring). They reference the Total row so they recompute in Excel.
    deduction_row = total_row + 2
    ws.cell(row=deduction_row, column=1, value="Check: Total Deductions")
    dcell = ws.cell(
        row=deduction_row, column=3, value=f"=(D{total_row}+E{total_row})"
    )
    dcell.number_format = MONEY_FORMAT

    reconcile_row = total_row + 3
    ws.cell(row=reconcile_row, column=1, value="Check: Gross - Deductions = Net")
    ws.cell(
        row=reconcile_row,
        column=3,
        value=(
            f"=IF(ROUND(C{total_row}-D{total_row}-E{total_row}-F{total_row},2)=0,"
            '"OK","MISMATCH")'
        ),
    )

    # Salary Disbursement Summary — the bank-upload table. Amount is a stored value
    # equal to that employee's Net Salary above (plain number, no "KES"): the bank
    # reads these raw cells and does not recalculate, so — like the detail numbers —
    # it must be a value, never a =F{row} reference. Comment is the fixed "salary".
    summary_title_row = reconcile_row + 2
    stitle = ws.cell(
        row=summary_title_row, column=1, value="Salary Disbursement Summary"
    )
    stitle.font = Font(bold=True)

    summary_header_row = summary_title_row + 1
    for col, header in ((1, "Phone Number"), (2, "Amount"), (3, "Comment")):
        cell = ws.cell(row=summary_header_row, column=col, value=header)
        cell.font = Font(bold=True)

    summary_row = summary_header_row + 1
    for phone, net in summary_entries:
        ws.cell(row=summary_row, column=1, value=phone)
        amount_cell = ws.cell(row=summary_row, column=2, value=float(net))
        amount_cell.number_format = PLAIN_AMOUNT_FORMAT
        ws.cell(row=summary_row, column=3, value="salary")
        summary_row += 1

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
